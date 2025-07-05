# -*- coding: utf-8 -*-
"""
Created on Fri Aug  2 08:58:50 2024

@author: chanboonkwee
"""
import os
import pandas as pd
import numpy as np
import openpyxl
import traceback
from tools.inputmapper import InputMapper
# from tools.misc import update_spreadsheet
from tools.xwtools import set_sensitivity_label

class Logger:
    fqfn='http://Excel.logfile.residing.in/on-prem.sharepoint.accessible.from.local/logfile.xlsx'
    def __init__(self, fileurl:str='', stdout=print, **kwargs):
        if fileurl == '':
            fileurl = self.fqfn

        url, filename = os.path.split(fileurl)
        if not os.path.exists(url):
            if url == '':
                raise FileNotFoundError(f"URL: '{url}' FILE: '{filename}'")

        self.obj = InputMapper(cell=filename,
                               file_pattern=filename,
                               url=url,
                               date_fmt='',
                               pattern='.*',
                               verbose=False,
                               stdout=stdout)

        self.fullpath_filename = self.obj.fullpath_filename
        df = pd.read_excel(self.fullpath_filename, sheet_name='Sheet1')
        kwargs['Log_ID'] = len(df) + 1
        # print(list(kwargs.keys()))
        df = df.append(kwargs, ignore_index=True)
        # print(df)
        # append_log_xlsx(filename=self.fullpath_filename, _df=df, sheet_name='Sheet1')
        update_spreadsheet(path=self.fullpath_filename, _df=df,
                            startrow=2, sheet_name='Sheet1')
        # set_sensitivity_label(self.fullpath_filename)

    def __del__(self):
        if hasattr(self, 'obj'):
            del self.obj


def update_spreadsheet(path:str ='',
                       _df=None,
                       startcol:int=1,
                       startrow:int=1,
                       sheet_name:str ="Sheet1",
                       clear:bool=False):
    if not os.path.exists(path):
        raise FileNotFoundError(f'{path} not found.')
    if _df is None:
        tb = traceback.format_exc()
        raise ValueError('No data').with_traceback(tb)
    wb = openpyxl.load_workbook(path)
    if sheet_name not in wb.sheetnames:
        tb = traceback.format_exc()
        raise ValueError(f'<Sheet \'{sheet_name}\'> missing in {path}')
    ws = wb[sheet_name]

    # Clear all content in the worksheet
    # ws.delete_rows(1, ws.max_row)

    # Clear all content in the worksheet except for the header row
    try:
        if clear:
            for row in ws:
                for cell in row:
                    cell.value = None
    except AttributeError:
        print(f"AttributeError at ({row}, {cell})")
        raise
    # if clear:
    #     for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    #         for cell in row:
    #             try:
    #                 cell.value = None
    #             except AttributeError:
    #                 continue

    for (row, col), cell_value in np.ndenumerate(_df.values):
        try:
            if pd.isna(cell_value):
                ws.cell(row=startrow + row, column=startcol + col).value = None
            else:
                ws.cell(row=startrow + row, column=startcol + col).value = cell_value
        except AttributeError:
            print(f"Attribute Error encountered r:{startrow+row} c:{startcol+col}")
            raise
    tb_name = [i for i in ws._tables.keys()][0]
    tb = ws._tables[tb_name]
    tb.ref = f"A1:{openpyxl.utils.get_column_letter(_df.shape[1])}{_df.shape[0] + 1}"
    # for row in range(0, _df.shape[0]): #For each row in the dataframe
    #     for col in range(0, _df.shape[1]): #For each column in the dataframe
    #         cell_value = _df.iat[row, col]
    #         try:
    #             if pd.isna(cell_value):
    #                 ws.cell(row = startrow + row, column = startcol + col).value = None
    #             else:
    #                 ws.cell(row = startrow + row, column = startcol + col).value = cell_value
    #         except AttributeError:
    #             print(f"Attribute Error encountered r:{startrow+row} c:{startcol+col}")
    #             raise
    wb.save(path)
    wb.close()
    del wb



if __name__=='__main__':
    obj = Logger(Script_Name='DRB Dashboard Filelist',
                 Script_Version='1.0b',
                 Start_Date='',
                 Start_Time='',
                 UserID='chanboonkwee',
                 End_Date='',
                 End_Time='',
                 Status='',
                 Exception='Nil')
    print(obj.fullpath_filename)
    # df_dict = pd.read_excel(obj.fullpath_filename, sheet_name=None)
    # df = df_dict['Sheet1']
    # print(df.columns)
    # print(f"{len(df)} rows")
