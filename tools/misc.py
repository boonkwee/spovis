"""
Created on

@author: chanboonkwee
"""
import re
import ctypes
import itertools
import string
import os
import random
import math
import openpyxl
import pandas as pd
import numpy as np
import datetime
import subprocess
import traceback
from urllib.parse import quote
import urllib.parse as parse
from pandas import DataFrame
from decimal import Decimal
# https://stackoverflow.com/questions/1350671/inner-exception-with-traceback-in-python
import tkinter as tk
from tkinter import filedialog
from types import GeneratorType


def read_csv_lines(file_path:str='', x:int=0, encoding:str='utf-8')-> list:
    data = []
    with open(file_path, 'r', encoding=encoding) as file:
        for _ in range(x):
            line = file.readline().strip()
            if not line:
                break
            data.append(line.split(','))
    return data


def add_new_data(df1, df2, common_column):
    # Merge the dataframes on the common column
    merged_df = pd.merge(df1, df2, on=common_column, how='outer', suffixes=('_A', '_B'))

    # Define a custom function to handle the merging logic
    def merge_logic(row):
        for col in df1.columns:
            if col + '_A' in row.keys() and col + '_B' in row.keys():
                if row[col + '_A'] == row[col + '_B']:
                    row[col] = row[col + '_B']
                else:
                    row[col] = row[col + '_A']
        return row

    # Apply the custom function to each row of the merged dataframe
    merged_df = merged_df.apply(merge_logic, axis=1)

    # Drop the redundant columns
    list1 = [col + '_A' for col in df1.columns] + [col + '_B' for col in df2.columns]
    list2 = merged_df.columns.to_list()
    merged_df = merged_df.drop(columns=list(set(list1) & set(list2)))

    return merged_df


def select_xlsx_file(**kwargs):
    root = tk.Tk()
    root.wm_attributes('-topmost', 1)
    root.withdraw()  # Hide the main window

    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")],
                                           **kwargs)

    return file_path


class as_local:
    def __init__(self, path: str ='', verbose: bool=False, stdout=print):
        '''
        Given a url, will attempt to map to a local drive that is available.
        Once mapped successfully, will set self.drive_status to zero
        Parameters
        ----------
        path : str, optional
            DESCRIPTION. The default is ''.
        verbose : bool, optional
            DESCRIPTION. The default is False.

        Raises
        ------
        IOError
            DESCRIPTION.
        ValueError
            DESCRIPTION.

        Returns
        -------
        None.

        '''
        self.raw_path = to_raw(path)
        self.drive_status = -1
        self.verbose = verbose
        self.drive = ''
        self.stdout = stdout

        self.filepath, self.file = os.path.split(self.raw_path)
        if is_url(self.filepath):
            self.drive = random_available_drive()
            if self.drive == '':
                raise IndexError('No more drive letters available. Please disconnect mapped drives to continue')
            self.drive_status = map_drive(self.drive, self.filepath)
            if self.drive_status > 0:
                raise IOError(f"Unable to map to sharepoint to access '{self.file}'")
            if verbose:
                self.stdout(f'{self.filepath} mapped as {self.drive}:')
        else:
            if os.path.exists(self.filepath):
                self.drive = os.path.splitdrive(self.filepath)[0]
            else:
                raise ValueError(f'Invalid path string: [{path}]')
        # self.path = os.path.join(f'{self.drive}:', os.sep, self.file)

    @property
    def path(self) -> str:
        """
        Returns
        -------
        str
            Returns the mapped, or local path without the filename.

        """
        return self.filepath if os.path.exists(self.filepath) else os.path.join(f'{self.drive}:', os.sep)


    @property
    def fullpath(self) -> str:
        """
        Returns
        -------
        str
            Returns the mapped, or local path with the filename

        """
        return parse.unquote(os.path.join(f'{self.drive}:', os.sep, self.file))


    def __del__(self):
        if self.drive_status == 0:
            if self.drive in drives_in_use():
                status = disconnect_drive(self.drive)
                if status != 0:
                    if self.verbose:
                        self.stdout(f'Disconnect {self.drive}: failed')
                else:
                    self.stdout(f'Disconnect {self.drive}: successful')
                    self.drive_status = -1
                    self.drive = ''


def ex_months(start_dt:datetime=datetime.date.today(),
              end_dt:datetime=datetime.date.today())-> list:
    """
    Return the months involved between 2 dates in a list

    The following example illustrates returning list of [1,2,3] when given start date as 1 Jan
    and end date as 3 Mar. the return value is [Jan, Feb Mar] in numerical values

    e.g. ex_months(datetime(2023,1,3), datetime(2023,3,3)) -> [1,2,3]

    Parameters
    ----------
    start_dt : datetime, optional
        Start date. The default is today.
    end_dt : datetime, optional
        End date. The default is today.

    Returns
    -------
    list
        of months involved in numerical value.

    """
    start =  start_dt.month
    # today = datetime.date.today()
    end = end_dt.month
    return list(range(start, end+1))


def num_monthname(num:int)-> str:
    """
    Return the name of month, given the corresponding numerical month value.
    If the corresponding numerical month is out of range, e.g. 0 or 13, returns empty string ''
    e.g. num_monthname(1) -> 'January'

    Parameters
    ----------
    num : int
        numerical value of month, 1 - January, 12 December.

    Returns
    -------
    str
        Name of month.

    """
    t = datetime.date.today()
    return datetime.date(t.year, num, 1).strftime('%B') if 0 < num <= 12 else ''


def unc_to_url(full_path: str='', protocol: str='https:')-> str:
    """
    Converts UNC pathname to url, will convert spaces to %20
    Works specifically for sharepoint-UNCs. As sharepoint UNC will contain some_url@server_site

    Parameters
    ----------
    full_path : str, optional
        DESCRIPTION. The default is ''.

    Returns
    -------
    str
        DESCRIPTION.

    """
    if full_path == '':
        return ''
    path = full_path
    pack = []
    loop = True
    while loop:
        head,tail = os.path.split(path)
        path = head
        if tail != '':
            pack.append(quote(tail))
        else:
            if head.endswith(os.sep):
                path = head.rstrip(os.sep)
            if head == path:
                index_at = head.find('@')
                index_slash = head.find(os.sep, index_at)
                result = head[:index_at] + head[index_slash:]
                pack.append(result)
                loop = False
    # domain_name = os.path.join(protocol)
    # print(domain_name)
    full_url = protocol + os.sep.join(reversed(pack)).replace(os.sep, '/')

    return full_url if full_url.endswith('/') else full_url + '/'


def parse_net_use() -> dict:
    """
    Runs 'net use', parse the output and returns the output in a dictionary
    NET USE is run to discover the linking url for mapped drives.


    Returns
    -------
    dict
        DESCRIPTION.

    """
    net_use_output = subprocess.check_output(['net', 'use']).decode('utf-8')
    ps = net_use_output.split('\r\n')
    end_lines = [
        'There are no entries in the list',
        'The command completed successfully']
    # print(f"{net_use_output}")
    end_flag = False
    net_use_dict = {}
    for line_count, p in enumerate(ps):
        for l in end_lines:
            if l in p:
                end_flag = True
        if line_count < 6 or end_flag:
            continue
        many_parts = p.split()
        part1 = many_parts[0]
        part2 = ' '.join(many_parts[1:])
        if len(part1) > 2:
            continue
        # print(f"'{part1} ({len(part1)})', '{part2}'")
        net_use_dict[part1] = part2
    return net_use_dict


def disconnect_all_net_drive(drive_list: list=None, verbose:bool=True, reply_to=print):
    '''
    Use the Microsoft "NET USE" command through the os module to disconnect all
    network drives. Refer to the disconnect_drive function for more details.

    Parameters
    ----------
    drive_list : list, optional
        leave blank to disconnect all network drives.
        Provide a list e.g.

            disconnect_all_net_drive(['A', 'B'])

        to disconnect A and B drive, disconnect_all_net_drive(['A:', 'B:'])
        does the same
    verbose : bool, optional
        Displays all message when disconnecting drives. The default is True.
    Returns
    -------
    None.

    '''
    if drive_list is None:
        drive_list = list(parse_net_use().keys())
    if len(drive_list) < 1:
        if verbose:
            msg ='No drive to disconnect'
            reply_to(msg)
    for drive in drive_list:
        if verbose:
            msg = f'Disconnecting {drive}'
            reply_to(msg)
        try:
            disconnect_drive(drive[0])
            msg = '.. done'
            if verbose:
                reply_to(msg)
        except Exception as ex:
            msg = f'.. failed ({type(ex).__name__})'
            if verbose:
                reply_to(msg)


def mapped_urls():
    mapped_dict = {key: unc_to_url(value) for key, value in parse_net_use().items()}
    return mapped_dict

def to_raw(input_string : str='') -> str:
    '''
    Parameters
    ----------
    input_string : str, optional
        DESCRIPTION. The default is ''.

    Returns
    -------
    str
        raw format of string.

    '''
    return r"{}".format(input_string)


def clean_column_spaces(df:DataFrame) -> DataFrame:
    '''
    Parameters
    ----------
    df : DataFrame
        DESCRIPTION.

    Returns
    -------
    DataFrame
        same DataFrame with multiple-contiguous spaces in columns removed.
    '''
    copy_df = df.copy()
    col_header = copy_df.columns.to_list()
    copy_df.columns = [(' '.join(c.split()) if not pd.isna(c) else c) for c in col_header]
    return copy_df


def clean_dict_key(original_dict:dict):
    new_dict = {}
    for k, v in original_dict.items():
        new_k = ' '.join(k.split())
        new_dict[new_k] = v
    return new_dict


# Function to truncate to 3 decimal places with error handling
def truncate_ndecimal(value, xdp:int=3):
    if isinstance(value, str):
        value = float(value.strip())
    try:
        return round(float(value), xdp)
    except (ValueError, TypeError):
        return value  # Handle invalid values as per your requirement


def clean_invisible_char(s:str=''):
    return re.sub(r'[^\x20-\x7E]', '', s)


def update_spreadsheet(path:str ='',
                       _df=None,
                       startcol:int=1,
                       startrow:int=1,
                       sheet_name:str ="Sheet1",
                       clear:bool=False):
    if not os.path.exists(path):
        raise FileNotFoundError(f'{path} not found.')
    if _df is None:
        tb = traceback.format_exc()
        raise ValueError('No data').with_traceback(tb)
    wb = openpyxl.load_workbook(path, data_only=True)
    if sheet_name not in wb.sheetnames:
        tb = traceback.format_exc()
        raise ValueError(f'<Sheet \'{sheet_name}\'> missing in {path}')
    ws = wb[sheet_name]

    # Clear all content in the worksheet
    # ws.delete_rows(1, ws.max_row)

    # Clear all content in the worksheet except for the header row
    try:
        if clear:
            for row in ws:
                for cell in row:
                    cell.value = ''
    except AttributeError:
        print(f"AttributeError at ({row}, {cell})")
        raise
    # if clear:
    #     for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    #         for cell in row:
    #             try:
    #                 cell.value = None
    #             except AttributeError:
    #                 continue

    for (row, col), cell_value in np.ndenumerate(_df.values):
        try:
            if pd.isna(cell_value):
                ws.cell(row=startrow + row, column=startcol + col).value = ''
            else:
                ws.cell(row=startrow + row, column=startcol + col).value = cell_value
        except AttributeError:
            print(f"Attribute Error encountered r:{startrow+row} c:{startcol+col}")
            raise
    # for row in range(0, _df.shape[0]): #For each row in the dataframe
    #     for col in range(0, _df.shape[1]): #For each column in the dataframe
    #         cell_value = _df.iat[row, col]
    #         try:
    #             if pd.isna(cell_value):
    #                 ws.cell(row = startrow + row, column = startcol + col).value = ''
    #             else:
    #                 ws.cell(row = startrow + row, column = startcol + col).value = cell_value
    #         except AttributeError:
    #             print(f"Attribute Error encountered r:{startrow+row} c:{startcol+col}")
    #             raise

    wb.save(path)
    wb.close()
    del wb


def update_spreadsheet_format(path:str       ='',
                              _df            = None,
                              startcol:int   = 1,
                              startrow:int   = 1,
                              sheet_name:str = "Sheet1",
                              _fmt:dict      = None,
                              clear:bool     = False):
    if not os.path.exists(path):
        raise FileNotFoundError(f'{path} not found.')
    if _df is None:
        raise ValueError('No data')

    wb = openpyxl.load_workbook(path)
    if sheet_name not in wb.sheetnames:
        tb = traceback.format_exc()
        raise ValueError(f'<Sheet \'{sheet_name}\'> missing in {path}').with_traceback(tb)
    ws = wb[sheet_name]
    for row in range(0, _df.shape[0]): #For each row in the dataframe
        for col in range(0, _df.shape[1]): #For each column in the dataframe
            cell_value = _df.iat[row, col]
            if pd.isna(cell_value):
                if clear:
                    ws.cell(row = startrow + row, column = startcol + col).value = ''
            else:
                ws.cell(row = startrow + row, column = startcol + col).value = cell_value
            if _fmt:
                for k, fmt in _fmt.items():
                    if _df.columns[col] == k:
                        ws.cell(row = startrow + row, column = startcol + col).number_format = fmt

    wb.save(path)
    wb.close()
    del wb


def eq_pyxl(python_value, excel_value, epsilon=Decimal('1e-9')) -> bool:
    if math.isnan(python_value) or math.isnan(excel_value):
        return False
    # Convert values to Decimal for precision
    python_decimal = Decimal(str(python_value))
    excel_decimal = Decimal(str(excel_value))

    # Compare using tolerance
    are_equal = abs(python_decimal - excel_decimal) < epsilon

    return are_equal


def get_missing_items(compare_list: list=None, reference_list: list=None) -> list:
    """
    Parameters
    ----------
    compare_list : list, optional
        compared against the reference_list. The default is None.
    reference_list : list, optional
        Reference list where the compare_list is compared against. The default is None.

    Returns
    -------
    list
        empty list if reference list is same as compare_list, otherwise returns
        the list of items from reference list that is not found in compare_list.
    Example
    get_missing_items(compare_list=['A', 'B', 'C'], reference_list=['A', 'B', 'C', 'D']) returns ['D']

    """
    if compare_list is None or reference_list is None:
        return None
    missing_item = [item not in compare_list for item in reference_list]
    return [item for item, flag in zip(reference_list, missing_item) if flag]


def drives_in_use() -> list:
    """
    Returns
    -------
    list
        of drive letters that is NOT available for mapping
    """
    if os.name != 'nt':
        return []
    drive_bitmask = ctypes.cdll.kernel32.GetLogicalDrives()
    return list(itertools.compress(string.ascii_uppercase,
            map(lambda x:ord(x) - ord('0'), bin(drive_bitmask)[:1:-1])))


def drives_not_in_use() -> list:
    """
    Returns
    -------
    list
        of drive letters that is available for mapping
    """
    if os.name != 'nt':
        return []
    drive_bitmask = ctypes.cdll.kernel32.GetLogicalDrives()
    in_use = list(itertools.compress(string.ascii_uppercase,
            map(lambda x:ord(x) - ord('0'), bin(drive_bitmask)[:1:-1])))
    return [letter for letter in string.ascii_uppercase if not letter in in_use]


def map_drive(drive_letter: str, url: str, persistent: bool=False) -> int:
    """
    Parameters
    ----------
    drive_letter : str
        A alphabet letter to map drive to.
    url : str
        The url that will be map to the drive_letter

    Raises
    ------
    ValueError
        if url is invalid
    IOError
        If drive_letter is already in use

    Returns
    -------
    int
        0 if successfully mapped, > 0 if mapping failed.
    """
    if not is_url(url):
        tb = traceback.format_exc()
        raise ValueError('Attempting to map invalid url').with_traceback(tb)
    drive = drive_letter[:1].upper()
    if drive in drives_in_use():
        tb = traceback.format_exc()
        raise IOError(f'{drive_letter} already in use').with_traceback(tb)

    value = os.system(f'net use {drive}: {url} {"/persistent:yes" if persistent else "/persistent:yes"}')
    return value
    # if value == 0:
    #     print(f'{drive}: is now mapped')
    # else:
    #     print('Map drive failed')


def disconnect_drive(drive_letter: str) -> int:
    """
    Parameters
    ----------
    drive_letter : str
        A alphabet letter indicating the drive to disconnect.

    Returns
    -------
    int
        0 if successfully mapped, != 0 if mapping failed.
    """
    drive = drive_letter[0].upper()
    if os.path.ismount(f'{drive}:'):
        return os.system(f'net use {drive}: /delete')
    return 2

def random_available_drive() -> str:
    """
    Returns
    -------
    str
        A alphabet letter indicating a drive that is available for mapping
    """
    if os.name != 'nt':
        return []
    drive_bitmask = ctypes.cdll.kernel32.GetLogicalDrives()
    in_use = list(itertools.compress(string.ascii_uppercase,
            map(lambda x:ord(x) - ord('0'), bin(drive_bitmask)[:1:-1])))
    try:
        assigned_letter = random.choice([letter for letter in string.ascii_uppercase if not letter in in_use][:5])
    except IndexError as e:
        tb = traceback.format_exc()
        raise e('No more drive letters available, please disconnect mapped drives to continue').with_traceback(tb)
    return assigned_letter


def is_url(url_str: str) -> bool:
    url_pattern = r'https?:\/\/(www\.)?[-a-zA-Z0-9@:%._\+~#=]{1,256}\.[a-zA-Z0-9()]{1,6}\b([-a-zA-Z0-9()@:%_\+.~#?&//=]*)'
    url_regex = re.compile(url_pattern, re.IGNORECASE)
    return re.match(url_regex, url_str) is not None


def list_all_cols(mylist: list=[]) -> str:
    if isinstance(mylist, GeneratorType):
        mylist = list(mylist)
    list_size = len(mylist)
    if list_size < 1:
        return ''
    retn_str = ''
    char_cnt = math.ceil(math.log10(list_size))
    # char_item = math.ceil(math.log10(max(len(i) for i in mylist)))
    str_fmt = f"%{char_cnt}d %-s\n"
    # print(str_fmt)
    for index, row in enumerate(mylist):
        retn_str += str_fmt % (index+1, row)
    return retn_str


def new_tempfile(filename:str='', length:int=10):
    path, basefile = os.path.split(filename)
    fn, ext = os.path.splitext(basefile)
    letters = string.ascii_letters
    new_basefile = ''.join(random.choice(letters) for _ in range(length)) + ext
    return os.path.join(path, new_basefile)


if __name__ == '__main__':
    print(list_all_cols([str(i) for i in range(101)]))
    # Example usage
    random_text = new_tempfile(10, 'c:\\some_folder\\new.txt')
    print(random_text)

